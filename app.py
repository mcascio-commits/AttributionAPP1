import os, json, shutil, smtplib, threading
from datetime import datetime
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import (Flask, render_template, request, jsonify, redirect,
                   url_for, send_file, flash, session)
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from database import get_db, fetchall, fetchone, execute, lastid, init_db, seed

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production-xyz987')

# ── Flask-Login ───────────────────────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter pour accéder à cette page.'

class User(UserMixin):
    def __init__(self, id, username, role, actif):
        self.id       = id
        self.username = username
        self.role     = role
        self.actif    = actif

    @property
    def is_admin(self):
        return self.role == 'admin'

    @property
    def can_edit(self):
        return self.role == 'admin'

@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    u = fetchone(conn, "SELECT * FROM utilisateurs WHERE id=?", (int(user_id),))
    conn.close()
    if u and u['actif']:
        return User(u['id'], u['username'], u['role'], u['actif'])
    return None

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            return jsonify({'ok': False, 'error': 'Accès refusé — compte administrateur requis'}), 403
        return f(*args, **kwargs)
    return decorated

# ── Helpers ───────────────────────────────────────────────────────────────────
def annee_active():
    conn = get_db()
    r = fetchone(conn, "SELECT label FROM annees WHERE actif=1")
    conn.close()
    return r['label'] if r else '2025-2026'

def nav_data():
    conn = get_db()
    filieres = fetchall(conn, "SELECT * FROM filieres WHERE actif=1 ORDER BY ordre")
    annees   = fetchall(conn, "SELECT * FROM annees ORDER BY label DESC")
    conn.close()
    return filieres, annees, annee_active()

def _ntpp_total(annee):
    conn = get_db()
    cats = fetchall(conn, """
        SELECT c.signe, COALESCE(v.valeur,0) as valeur
        FROM ntpp_categories c
        LEFT JOIN ntpp_valeurs v ON v.categorie_id=c.id AND v.annee=?
        WHERE c.parent_id IS NULL
    """, (annee,))
    sous = fetchone(conn, """
        SELECT COALESCE(SUM(v.valeur),0) as total
        FROM ntpp_categories c
        LEFT JOIN ntpp_valeurs v ON v.categorie_id=c.id AND v.annee=?
        WHERE c.parent_id IS NOT NULL
    """, (annee,))
    conn.close()
    total = sum(c['signe'] * c['valeur'] for c in cats)
    total += sous['total'] if sous else 0
    return total

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        conn = get_db()
        u = fetchone(conn, "SELECT * FROM utilisateurs WHERE username=? AND actif=1", (username,))
        conn.close()
        if u and check_password_hash(u['password'], password):
            user = User(u['id'], u['username'], u['role'], u['actif'])
            login_user(user, remember=True)
            return redirect(request.args.get('next') or url_for('index'))
        flash('Identifiants incorrects', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ── Pages ─────────────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    filieres, annees, annee = nav_data()
    return render_template('index.html', filieres=filieres, annees=annees, annee=annee)

@app.route('/filiere/<int:fid>')
@login_required
def filiere(fid):
    annee = request.args.get('annee', annee_active())
    conn  = get_db()
    fil   = fetchone(conn, "SELECT * FROM filieres WHERE id=?", (fid,))
    if not fil: conn.close(); return redirect('/')
    filieres = fetchall(conn, "SELECT * FROM filieres WHERE actif=1 ORDER BY ordre")
    annees   = fetchall(conn, "SELECT * FROM annees ORDER BY label DESC")
    classes  = fetchall(conn, "SELECT * FROM classes WHERE filiere_id=? ORDER BY ordre", (fid,))
    cours    = fetchall(conn, "SELECT * FROM cours WHERE filiere_id=? ORDER BY type,ordre", (fid,))
    attrs    = fetchall(conn, """
        SELECT a.*, p.acronyme, p.prenom, p.nom as pnom,
               c.nom as cours_nom, c.heures as cours_heures, c.type as cours_type
        FROM attributions a
        JOIN personnel p ON a.personnel_id=p.id
        JOIN cours c ON a.cours_id=c.id
        WHERE c.filiere_id=? AND a.annee=?
    """, (fid, annee))
    titulaires = fetchall(conn, """
        SELECT t.classe_id, p.acronyme, p.prenom, p.nom as pnom, t.id
        FROM titulaires t JOIN personnel p ON t.personnel_id=p.id
        WHERE t.annee=? AND t.classe_id IN (SELECT id FROM classes WHERE filiere_id=?)
    """, (annee, fid))
    coord_cats = fetchall(conn, "SELECT * FROM coord_categories ORDER BY ordre")
    eleves     = fetchall(conn, """
        SELECT * FROM eleves_options WHERE cours_id IN
        (SELECT id FROM cours WHERE filiere_id=?) AND annee=?
    """, (fid, annee))
    ntpp_total = _ntpp_total(annee)
    h_total = fetchone(conn, """
        SELECT COALESCE(SUM(CASE WHEN a.heures_attr IS NOT NULL THEN a.heures_attr ELSE c.heures END),0) as h
        FROM attributions a JOIN cours c ON a.cours_id=c.id WHERE a.annee=?
    """, (annee,))['h']
    conn.close()
    return render_template('filiere.html',
        fil=fil, filieres=filieres, annees=annees, annee=annee,
        classes=classes, cours=cours, attrs=attrs, titulaires=titulaires,
        coord_cats=coord_cats, eleves=eleves,
        ntpp_total=ntpp_total, h_total=h_total)

@app.route('/ntpp')
@login_required
def ntpp():
    annee = request.args.get('annee', annee_active())
    conn  = get_db()
    filieres = fetchall(conn, "SELECT * FROM filieres WHERE actif=1 ORDER BY ordre")
    annees   = fetchall(conn, "SELECT * FROM annees ORDER BY label DESC")
    cats = fetchall(conn, """
        SELECT c.*, COALESCE(v.valeur,0) as valeur, v.id as vid
        FROM ntpp_categories c
        LEFT JOIN ntpp_valeurs v ON v.categorie_id=c.id AND v.annee=?
        ORDER BY CASE WHEN c.parent_id IS NULL THEN 0 ELSE 1 END, c.parent_id, c.ordre
    """, (annee,))
    h_total = fetchone(conn, """
        SELECT COALESCE(SUM(CASE WHEN a.heures_attr IS NOT NULL THEN a.heures_attr ELSE c.heures END),0) as h
        FROM attributions a JOIN cours c ON a.cours_id=c.id WHERE a.annee=?
    """, (annee,))['h']
    conn.close()
    return render_template('ntpp.html', filieres=filieres, annees=annees, annee=annee,
                           cats=cats, ntpp_total=_ntpp_total(annee), h_total=h_total)

@app.route('/personnel')
@login_required
def personnel():
    filieres, annees, annee = nav_data()
    conn = get_db()
    people = fetchall(conn, "SELECT * FROM personnel ORDER BY nom, prenom")
    conn.close()
    return render_template('personnel.html', filieres=filieres, annees=annees,
                           annee=annee, people=people)

@app.route('/synthese')
@login_required
def synthese():
    annee = request.args.get('annee', annee_active())
    filieres, annees, _ = nav_data()
    conn = get_db()
    data = fetchall(conn, """
        SELECT p.id, p.acronyme, p.prenom, p.nom, p.email, p.heures_min, p.heures_max,
               COALESCE(SUM(CASE WHEN a.heures_attr IS NOT NULL THEN a.heures_attr ELSE c.heures END),0) as total,
               COALESCE((SELECT SUM(n.heures) FROM nominations n WHERE n.personnel_id=p.id),0) as total_nomination
        FROM personnel p
        LEFT JOIN attributions a ON a.personnel_id=p.id AND a.annee=?
        LEFT JOIN cours c ON a.cours_id=c.id
        WHERE p.actif=1 GROUP BY p.id,p.acronyme,p.prenom,p.nom,p.email,p.heures_min,p.heures_max
        ORDER BY p.nom, p.prenom
    """, (annee,))
    titu = fetchall(conn, """
        SELECT t.personnel_id, f.nom as filiere, cl.nom as classe
        FROM titulaires t JOIN classes cl ON t.classe_id=cl.id
        JOIN filieres f ON cl.filiere_id=f.id WHERE t.annee=?
    """, (annee,))
    conn.close()
    return render_template('synthese.html', filieres=filieres, annees=annees,
                           annee=annee, data=data, titu=titu)

@app.route('/recap')
@login_required
def recap():
    annee = request.args.get('annee', annee_active())
    conn  = get_db()
    filieres_list = fetchall(conn, "SELECT * FROM filieres WHERE actif=1 ORDER BY ordre")
    annees_list   = fetchall(conn, "SELECT * FROM annees ORDER BY label DESC")
    recap_data = []
    for fil in filieres_list:
        cours_list = fetchall(conn, "SELECT * FROM cours WHERE filiere_id=? ORDER BY type,ordre", (fil['id'],))
        fil_vides = 0; cours_recap = []
        for c in cours_list:
            nb_grp = c['nb_groupes'] or 1
            attr_groups = fetchall(conn, "SELECT DISTINCT groupe_num FROM attributions WHERE cours_id=? AND annee=?", (c['id'], annee))
            attr_set = {a['groupe_num'] for a in attr_groups}
            groupes_vides = [g for g in range(1, nb_grp+1) if g not in attr_set]
            elv = fetchone(conn, "SELECT nb_eleves FROM eleves_options WHERE cours_id=? AND annee=?", (c['id'], annee))
            if groupes_vides:
                fil_vides += len(groupes_vides)
                cours_recap.append({'id':c['id'],'nom':c['nom'],'heures':c['heures'],
                    'type':c['type'],'nb_groupes':nb_grp,
                    'nb_attribues':nb_grp-len(groupes_vides),
                    'groupes_vides':groupes_vides,'nb_eleves':elv['nb_eleves'] if elv else None})
        recap_data.append({'id':fil['id'],'nom':fil['nom'],'degre':fil['degre'],
            'couleur':fil['couleur'],'nb_vides':fil_vides,'cours':cours_recap})
    conn.close()
    return render_template('recap.html', recap=recap_data,
                           filieres=filieres_list, annees=annees_list, annee=annee)

@app.route('/mails')
@login_required
def mails():
    filieres, annees, annee = nav_data()
    conn = get_db()
    people    = fetchall(conn, "SELECT * FROM personnel WHERE actif=1 AND email!='' ORDER BY nom")
    templates = fetchall(conn, "SELECT * FROM mail_templates")
    config    = fetchone(conn, "SELECT * FROM mail_config LIMIT 1")
    envois    = fetchall(conn, """
        SELECT e.*, p.acronyme, p.nom, p.prenom, t.nom as tnom
        FROM mail_envois e JOIN personnel p ON e.personnel_id=p.id
        LEFT JOIN mail_templates t ON e.template_id=t.id
        ORDER BY e.date_envoi DESC LIMIT 50
    """)
    conn.close()
    return render_template('mails.html', filieres=filieres, annees=annees, annee=annee,
                           people=people, templates=templates, config=config, envois=envois)

@app.route('/gestion')
@login_required
@admin_required
def gestion():
    filieres, annees, annee = nav_data()
    conn = get_db()
    all_filieres = fetchall(conn, "SELECT * FROM filieres ORDER BY ordre")
    conn.close()
    return render_template('gestion.html', filieres=filieres, all_filieres=all_filieres,
                           annees=annees, annee=annee)

@app.route('/utilisateurs')
@login_required
@admin_required
def utilisateurs():
    filieres, annees, annee = nav_data()
    conn = get_db()
    users = fetchall(conn, "SELECT id,username,role,actif,created_at FROM utilisateurs ORDER BY role,username")
    conn.close()
    return render_template('utilisateurs.html', filieres=filieres, annees=annees,
                           annee=annee, users=users)


# ── API Nominations ───────────────────────────────────────────────────────────
@app.route('/api/nominations/<int:pid>')
@login_required
def get_nominations(pid):
    conn = get_db()
    noms = fetchall(conn, "SELECT * FROM nominations WHERE personnel_id=? ORDER BY matiere", (pid,))
    conn.close()
    return jsonify(noms)

@app.route('/api/nomination', methods=['POST'])
@login_required
@admin_required
def add_nomination():
    d = request.json; conn = get_db()
    execute(conn, "INSERT INTO nominations(personnel_id,matiere,heures,type_cours) VALUES(?,?,?,?)",
            (d['personnel_id'], d['matiere'], d.get('heures',0), d.get('type_cours','FC')))
    conn.commit()
    nid = lastid(conn, 'nominations')
    conn.close()
    return jsonify({'ok':True,'id':nid})

@app.route('/api/nomination/<int:nid>', methods=['DELETE'])
@login_required
@admin_required
def del_nomination(nid):
    conn = get_db()
    execute(conn, "DELETE FROM nominations WHERE id=?", (nid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ── API Auth ──────────────────────────────────────────────────────────────────
@app.route('/api/utilisateur', methods=['POST'])
@login_required
@admin_required
def add_utilisateur():
    d = request.json
    if not d.get('username') or not d.get('password'):
        return jsonify({'ok':False,'error':'Username et mot de passe requis'})
    conn = get_db()
    try:
        execute(conn, "INSERT INTO utilisateurs(username,password,role) VALUES(?,?,?)",
                (d['username'].strip(), generate_password_hash(d['password']),
                 d.get('role','consultation')))
        conn.commit(); conn.close()
        return jsonify({'ok':True})
    except Exception as e:
        conn.close()
        return jsonify({'ok':False,'error':'Nom d\'utilisateur déjà pris'})

@app.route('/api/utilisateur/<int:uid>', methods=['PUT'])
@login_required
@admin_required
def update_utilisateur(uid):
    d = request.json; conn = get_db()
    if 'password' in d and d['password']:
        execute(conn, "UPDATE utilisateurs SET password=? WHERE id=?",
                (generate_password_hash(d['password']), uid))
    if 'role' in d:
        execute(conn, "UPDATE utilisateurs SET role=? WHERE id=?", (d['role'], uid))
    if 'actif' in d:
        execute(conn, "UPDATE utilisateurs SET actif=? WHERE id=?", (d['actif'], uid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/utilisateur/<int:uid>', methods=['DELETE'])
@login_required
@admin_required
def delete_utilisateur(uid):
    if uid == current_user.id:
        return jsonify({'ok':False,'error':'Impossible de supprimer votre propre compte'})
    conn = get_db()
    execute(conn, "DELETE FROM utilisateurs WHERE id=?", (uid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ── API Personnel ─────────────────────────────────────────────────────────────
@app.route('/api/personnel/search')
@login_required
def search_personnel():
    q = request.args.get('q','').strip().upper()
    if len(q) < 1: return jsonify([])
    conn = get_db()
    r = fetchall(conn, """
        SELECT id, acronyme, prenom, nom FROM personnel
        WHERE actif=1 AND (UPPER(acronyme) LIKE ? OR UPPER(nom) LIKE ? OR UPPER(prenom) LIKE ?)
        ORDER BY acronyme LIMIT 10
    """, (f'{q}%', f'{q}%', f'{q}%'))
    conn.close()
    return jsonify(r)

@app.route('/api/personnel', methods=['POST'])
@login_required
@admin_required
def add_personnel():
    d = request.json
    acro = d.get('acronyme','').strip().upper()
    if not acro: return jsonify({'ok':False,'error':'Acronyme requis'})
    conn = get_db()
    try:
        execute(conn, "INSERT INTO personnel(acronyme,prenom,nom,email,statut) VALUES(?,?,?,?,?)",
                (acro, d.get('prenom',''), d.get('nom',''), d.get('email',''), d.get('statut','')))
        conn.commit(); conn.close()
        return jsonify({'ok':True})
    except Exception as e:
        conn.close(); return jsonify({'ok':False,'error':str(e)})

@app.route('/api/personnel/<int:pid>', methods=['PUT'])
@login_required
@admin_required
def update_personnel(pid):
    d = request.json; conn = get_db()
    allowed = ('acronyme','prenom','nom','email','statut','actif','heures_min','heures_max')
    fields = [f"{k}=?" for k in d if k in allowed]
    vals   = [d[k] for k in d if k in allowed]
    if fields:
        vals.append(pid)
        execute(conn, f"UPDATE personnel SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/personnel/delete', methods=['POST'])
@login_required
@admin_required
def delete_personnel_bulk():
    ids = request.json.get('ids', [])
    if not ids: return jsonify({'ok':False,'error':'Aucun ID'})
    conn = get_db()
    for pid in ids:
        execute(conn, "DELETE FROM attributions WHERE personnel_id=?", (pid,))
        execute(conn, "DELETE FROM titulaires WHERE personnel_id=?", (pid,))
        execute(conn, "DELETE FROM personnel WHERE id=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True,'deleted':len(ids)})

@app.route('/api/personnel/import', methods=['POST'])
@login_required
@admin_required
def import_personnel():
    import pandas as pd, io as _io, csv
    f = request.files.get('file')
    if not f: return jsonify({'ok':False,'error':'Fichier manquant'})
    content = f.read()
    if f.filename.endswith(('.xlsx','.xls')):
        try:
            df = pd.read_excel(_io.BytesIO(content), header=0)
            cols = [str(c).strip() for c in df.columns]
            col_map = {}
            for i,c in enumerate(cols):
                cl = c.lower()
                if 'prénom' in cl or 'prenom' in cl: col_map['prenom'] = i
                elif 'nom' in cl: col_map['nom'] = i
                elif 'abrév' in cl or 'acronyme' in cl: col_map['acronyme'] = i
                elif 'mail' in cl or 'email' in cl: col_map['email'] = i
            data_rows = []
            for _, row_data in df.iterrows():
                r = {k: (str(row_data.iloc[i]).strip() if i < len(row_data) else '') for k,i in col_map.items()}
                r = {k: ('' if v in ('nan','None') else v) for k,v in r.items()}
                if r.get('acronyme'): data_rows.append(r)
        except Exception as e:
            return jsonify({'ok':False,'error':str(e)})
    else:
        text = content.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(_io.StringIO(text))
        data_rows = []
        for row_data in reader:
            acro = (row_data.get('acronyme') or row_data.get('Abréviation') or '').strip().upper()
            if acro:
                data_rows.append({'acronyme':acro,
                    'prenom':(row_data.get('prenom') or row_data.get('Prénom Enseignant') or '').strip(),
                    'nom':(row_data.get('nom') or row_data.get('Nom Enseignant') or '').strip(),
                    'email':(row_data.get('email') or row_data.get('mail prof :') or '').strip()})
    conn = get_db(); added = updated = 0
    for r in data_rows:
        acro = r.get('acronyme','').upper()
        if not acro: continue
        existing = fetchone(conn, "SELECT id,prenom,nom,email FROM personnel WHERE acronyme=?", (acro,))
        if existing:
            updates = {}
            if not existing['prenom'] and r.get('prenom'): updates['prenom'] = r['prenom']
            if not existing['nom'] and r.get('nom'):       updates['nom']    = r['nom']
            if not existing['email'] and r.get('email'):   updates['email']  = r['email']
            if updates:
                fields = [f"{k}=?" for k in updates]
                execute(conn, f"UPDATE personnel SET {','.join(fields)} WHERE acronyme=?",
                        list(updates.values()) + [acro])
                updated += 1
        else:
            execute(conn, "INSERT INTO personnel(acronyme,prenom,nom,email) VALUES(?,?,?,?)",
                    (acro, r.get('prenom',''), r.get('nom',''), r.get('email','')))
            added += 1
    conn.commit(); conn.close()
    return jsonify({'ok':True,'added':added,'updated':updated})

# ── API Attribution ───────────────────────────────────────────────────────────
@app.route('/api/attribution', methods=['POST'])
@login_required
@admin_required
def add_attribution():
    d = request.json; acro = d.get('acronyme','').strip().upper()
    conn = get_db()
    p = fetchone(conn, "SELECT id FROM personnel WHERE acronyme=?", (acro,))
    if not p:
        execute(conn, "INSERT INTO personnel(acronyme) VALUES(?)", (acro,))
        pid = lastid(conn, 'personnel')
    else: pid = p['id']
    try:
        execute(conn, """INSERT INTO attributions(cours_id,classe_id,personnel_id,annee,groupe_num,heures_attr)
                         VALUES(?,?,?,?,?,?)""",
                (d['cours_id'], d.get('classe_id'), pid,
                 d.get('annee', annee_active()), d.get('groupe_num',1), d.get('heures_attr')))
        conn.commit(); aid = lastid(conn, 'attributions')
        conn.close(); return jsonify({'ok':True,'id':aid})
    except Exception as e:
        conn.close(); return jsonify({'ok':False,'error':str(e)})

@app.route('/api/attribution/<int:aid>', methods=['PUT'])
@login_required
@admin_required
def update_attribution(aid):
    d = request.json; conn = get_db()
    if 'acronyme' in d:
        acro = d['acronyme'].strip().upper()
        p = fetchone(conn, "SELECT id FROM personnel WHERE acronyme=?", (acro,))
        if not p:
            execute(conn, "INSERT INTO personnel(acronyme) VALUES(?)", (acro,))
            pid = lastid(conn, 'personnel')
        else: pid = p['id']
        execute(conn, "UPDATE attributions SET personnel_id=? WHERE id=?", (pid, aid))
    if 'heures_attr' in d:
        execute(conn, "UPDATE attributions SET heures_attr=? WHERE id=?", (d['heures_attr'], aid))
    if 'couleur' in d:
        execute(conn, "UPDATE attributions SET couleur=? WHERE id=?", (d['couleur'] or None, aid))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/attribution/<int:aid>', methods=['DELETE'])
@login_required
@admin_required
def del_attribution(aid):
    conn = get_db()
    execute(conn, "DELETE FROM attributions WHERE id=?", (aid,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/attribution/swap', methods=['POST'])
@login_required
@admin_required
def swap_attributions():
    d = request.json; conn = get_db()
    a1 = fetchone(conn, "SELECT * FROM attributions WHERE id=?", (d['id1'],))
    a2 = fetchone(conn, "SELECT * FROM attributions WHERE id=?", (d['id2'],))
    if a1 and a2:
        execute(conn, "UPDATE attributions SET personnel_id=?,groupe_num=?,classe_id=? WHERE id=?",
                (a2['personnel_id'],a2['groupe_num'],a2['classe_id'],d['id1']))
        execute(conn, "UPDATE attributions SET personnel_id=?,groupe_num=?,classe_id=? WHERE id=?",
                (a1['personnel_id'],a1['groupe_num'],a1['classe_id'],d['id2']))
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/attribution/move', methods=['POST'])
@login_required
@admin_required
def move_attribution():
    d = request.json; conn = get_db()
    updates = []; vals = []
    for k in ('cours_id','groupe_num','classe_id'):
        if k in d: updates.append(f"{k}=?"); vals.append(d[k])
    if updates:
        vals.append(d['attr_id'])
        execute(conn, f"UPDATE attributions SET {','.join(updates)} WHERE id=?", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

# ── API Titulaires ────────────────────────────────────────────────────────────
@app.route('/api/titulaire', methods=['POST'])
@login_required
@admin_required
def add_titulaire():
    d = request.json; acro = d.get('acronyme','').strip().upper()
    conn = get_db()
    p = fetchone(conn, "SELECT id FROM personnel WHERE acronyme=?", (acro,))
    if not p:
        execute(conn, "INSERT INTO personnel(acronyme) VALUES(?)", (acro,))
        pid = lastid(conn, 'personnel')
    else: pid = p['id']
    try:
        execute(conn, "INSERT INTO titulaires(classe_id,personnel_id,annee) VALUES(?,?,?)",
                (d['classe_id'], pid, d.get('annee', annee_active())))
        conn.commit(); conn.close(); return jsonify({'ok':True})
    except Exception as e:
        conn.close(); return jsonify({'ok':False,'error':str(e)})

@app.route('/api/titulaire/<int:tid>', methods=['DELETE'])
@login_required
@admin_required
def del_titulaire(tid):
    conn = get_db()
    execute(conn, "DELETE FROM titulaires WHERE id=?", (tid,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

# ── API Cours ─────────────────────────────────────────────────────────────────
@app.route('/api/cours', methods=['POST'])
@login_required
@admin_required
def add_cours():
    d = request.json; conn = get_db()
    execute(conn, "INSERT INTO cours(filiere_id,nom,heures,type,ordre,nb_groupes) VALUES(?,?,?,?,?,?)",
            (d['filiere_id'],d['nom'],d.get('heures',0),d.get('type','FC'),d.get('ordre',999),d.get('nb_groupes',1)))
    conn.commit(); cid = lastid(conn,'cours'); conn.close()
    return jsonify({'ok':True,'id':cid})

@app.route('/api/cours/<int:cid>', methods=['PUT'])
@login_required
@admin_required
def update_cours(cid):
    d = request.json; conn = get_db()
    allowed = ('nom','heures','type','ordre','nb_groupes','coord_cat_id')
    fields = [f"{k}=?" for k in d if k in allowed]
    vals   = [d[k] for k in d if k in allowed]
    if fields:
        vals.append(cid)
        execute(conn, f"UPDATE cours SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/cours/<int:cid>', methods=['DELETE'])
@login_required
@admin_required
def del_cours(cid):
    conn = get_db()
    execute(conn, "DELETE FROM attributions WHERE cours_id=?", (cid,))
    execute(conn, "DELETE FROM eleves_options WHERE cours_id=?", (cid,))
    execute(conn, "DELETE FROM cours WHERE id=?", (cid,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/cours/reorder', methods=['POST'])
@login_required
@admin_required
def reorder_cours():
    items = request.json; conn = get_db()
    for item in items:
        execute(conn, "UPDATE cours SET ordre=? WHERE id=?", (item['ordre'], item['id']))
    conn.commit(); conn.close(); return jsonify({'ok':True})

# ── API Classes ───────────────────────────────────────────────────────────────
@app.route('/api/classe', methods=['POST'])
@login_required
@admin_required
def add_classe():
    d = request.json; conn = get_db()
    max_ordre = fetchone(conn, "SELECT COALESCE(MAX(ordre)+1,0) as o FROM classes WHERE filiere_id=?", (d['filiere_id'],))['o']
    execute(conn, "INSERT INTO classes(filiere_id,nom,ordre) VALUES(?,?,?)",
            (d['filiere_id'], d['nom'].strip().upper(), max_ordre))
    conn.commit(); cid = lastid(conn,'classes'); conn.close()
    return jsonify({'ok':True,'id':cid})

@app.route('/api/classe/<int:cid>', methods=['PUT'])
@login_required
@admin_required
def update_classe(cid):
    d = request.json; conn = get_db()
    if 'nom' in d:
        execute(conn, "UPDATE classes SET nom=? WHERE id=?", (d['nom'].strip().upper(), cid))
        conn.commit()
    if 'commentaire' in d:
        execute(conn, "UPDATE classes SET commentaire=? WHERE id=?", (d.get('commentaire') or None, cid))
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/classe/<int:cid>', methods=['DELETE'])
@login_required
@admin_required
def del_classe(cid):
    conn = get_db()
    execute(conn, "UPDATE attributions SET classe_id=NULL WHERE classe_id=?", (cid,))
    execute(conn, "DELETE FROM titulaires WHERE classe_id=?", (cid,))
    execute(conn, "DELETE FROM classes WHERE id=?", (cid,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/classes/<int:fid>')
@login_required
def get_classes(fid):
    conn = get_db()
    r = fetchall(conn, "SELECT * FROM classes WHERE filiere_id=? ORDER BY ordre", (fid,))
    conn.close(); return jsonify(r)

# ── API Élèves ────────────────────────────────────────────────────────────────
@app.route('/api/eleves', methods=['POST'])
@login_required
@admin_required
def save_eleves():
    d = request.json; annee = d.get('annee', annee_active()); conn = get_db()
    existing = fetchone(conn, "SELECT id,nb_eleves FROM eleves_options WHERE cours_id=? AND annee=?",
                        (d['cours_id'], annee))
    if existing:
        execute(conn, "UPDATE eleves_options SET nb_eleves=?,nb_eleves_precedent=?,source=? WHERE id=?",
                (d['nb_eleves'], existing['nb_eleves'], d.get('source','manuel'), existing['id']))
    else:
        execute(conn, "INSERT INTO eleves_options(cours_id,annee,nb_eleves,source) VALUES(?,?,?,?)",
                (d['cours_id'], annee, d['nb_eleves'], d.get('source','manuel')))
    conn.commit(); conn.close(); return jsonify({'ok':True})

# ── API NTPP ──────────────────────────────────────────────────────────────────
@app.route('/api/ntpp/categorie', methods=['POST'])
@login_required
@admin_required
def add_ntpp_cat():
    d = request.json or {}
    if not d.get('nom'): return jsonify({'ok':False,'error':'Nom requis'})
    conn = get_db()
    max_o = fetchone(conn, "SELECT COALESCE(MAX(ordre)+1,0) as o FROM ntpp_categories WHERE parent_id IS ?",
                     (d.get('parent_id'),))['o']
    execute(conn, "INSERT INTO ntpp_categories(nom,signe,ordre,parent_id) VALUES(?,?,?,?)",
            (d['nom'], d.get('signe',1), max_o, d.get('parent_id')))
    conn.commit(); cid = lastid(conn,'ntpp_categories'); conn.close()
    return jsonify({'ok':True,'id':cid})

@app.route('/api/ntpp/categorie/<int:cid>', methods=['PUT'])
@login_required
@admin_required
def update_ntpp_cat(cid):
    d = request.json; conn = get_db()
    fields = [f"{k}=?" for k in d if k in ('nom','signe','ordre')]
    vals   = [d[k] for k in d if k in ('nom','signe','ordre')]
    if fields:
        vals.append(cid)
        execute(conn, f"UPDATE ntpp_categories SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/ntpp/categorie/<int:cid>', methods=['DELETE'])
@login_required
@admin_required
def del_ntpp_cat(cid):
    conn = get_db()
    execute(conn, "DELETE FROM ntpp_valeurs WHERE categorie_id=?", (cid,))
    execute(conn, "DELETE FROM ntpp_categories WHERE id=? OR parent_id=?", (cid, cid))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/ntpp/valeur', methods=['POST'])
@login_required
@admin_required
def save_ntpp_val():
    d = request.json; annee = d.get('annee', annee_active()); conn = get_db()
    ex = fetchone(conn, "SELECT id FROM ntpp_valeurs WHERE categorie_id=? AND annee=?",
                  (d['categorie_id'], annee))
    if ex:
        execute(conn, "UPDATE ntpp_valeurs SET valeur=? WHERE id=?", (d['valeur'], ex['id']))
    else:
        execute(conn, "INSERT INTO ntpp_valeurs(categorie_id,annee,valeur) VALUES(?,?,?)",
                (d['categorie_id'], annee, d['valeur']))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/ntpp/reorder', methods=['POST'])
@login_required
@admin_required
def reorder_ntpp():
    items = request.json; conn = get_db()
    for item in items:
        execute(conn, "UPDATE ntpp_categories SET ordre=? WHERE id=?", (item['ordre'], item['id']))
    conn.commit(); conn.close(); return jsonify({'ok':True})

# ── API Coord ─────────────────────────────────────────────────────────────────
@app.route('/api/coord/categories')
@login_required
def get_coord_cats():
    conn = get_db()
    r = fetchall(conn, "SELECT * FROM coord_categories ORDER BY ordre")
    conn.close(); return jsonify(r)

@app.route('/api/coord/categorie', methods=['POST'])
@login_required
@admin_required
def add_coord_cat():
    d = request.json; conn = get_db()
    max_o = fetchone(conn, "SELECT COALESCE(MAX(ordre)+1,0) as o FROM coord_categories")['o']
    execute(conn, "INSERT INTO coord_categories(nom,couleur,ordre) VALUES(?,?,?)",
            (d['nom'], d.get('couleur','#888780'), max_o))
    conn.commit(); cid = lastid(conn,'coord_categories'); conn.close()
    return jsonify({'ok':True,'id':cid})

@app.route('/api/coord/categorie/<int:cid>', methods=['PUT'])
@login_required
@admin_required
def update_coord_cat(cid):
    d = request.json; conn = get_db()
    fields = [f"{k}=?" for k in d if k in ('nom','couleur','ordre')]
    vals   = [d[k] for k in d if k in ('nom','couleur','ordre')]
    if fields:
        vals.append(cid)
        execute(conn, f"UPDATE coord_categories SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/coord/categorie/<int:cid>', methods=['DELETE'])
@login_required
@admin_required
def del_coord_cat(cid):
    conn = get_db()
    execute(conn, "UPDATE cours SET coord_cat_id=NULL WHERE coord_cat_id=?", (cid,))
    execute(conn, "DELETE FROM coord_categories WHERE id=?", (cid,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

# ── API Filières ──────────────────────────────────────────────────────────────
@app.route('/api/filiere', methods=['POST'])
@login_required
@admin_required
def add_filiere():
    d = request.json; conn = get_db()
    max_o = fetchone(conn, "SELECT COALESCE(MAX(ordre)+1,0) as o FROM filieres")['o']
    execute(conn, "INSERT INTO filieres(nom,degre,couleur,ordre) VALUES(?,?,?,?)",
            (d['nom'], d.get('degre',''), d.get('couleur','#378ADD'), max_o))
    conn.commit(); fid = lastid(conn,'filieres')
    if d.get('copier_depuis'):
        for c in fetchall(conn, "SELECT * FROM cours WHERE filiere_id=?", (int(d['copier_depuis']),)):
            execute(conn, "INSERT INTO cours(filiere_id,nom,heures,type,ordre,nb_groupes) VALUES(?,?,?,?,?,?)",
                    (fid,c['nom'],c['heures'],c['type'],c['ordre'],c['nb_groupes']))
        conn.commit()
    conn.close(); return jsonify({'ok':True,'id':fid})

@app.route('/api/filiere/<int:fid>', methods=['PUT'])
@login_required
@admin_required
def update_filiere(fid):
    d = request.json; conn = get_db()
    fields = [f"{k}=?" for k in d if k in ('nom','degre','couleur','actif','ordre')]
    vals   = [d[k] for k in d if k in ('nom','degre','couleur','actif','ordre')]
    if fields:
        vals.append(fid)
        execute(conn, f"UPDATE filieres SET {','.join(fields)} WHERE id=?", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

# ── API Années ────────────────────────────────────────────────────────────────
@app.route('/api/annee', methods=['POST'])
@login_required
@admin_required
def set_annee():
    label = request.json.get('label'); conn = get_db()
    execute(conn, "UPDATE annees SET actif=0")
    if fetchone(conn, "SELECT id FROM annees WHERE label=?", (label,)):
        execute(conn, "UPDATE annees SET actif=1 WHERE label=?", (label,))
    else:
        execute(conn, "INSERT INTO annees(label,actif) VALUES(?,1)", (label,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/annee/nouvelle', methods=['POST'])
@login_required
@admin_required
def nouvelle_annee():
    d = request.json; label = d['label']; conn = get_db()
    try:
        execute(conn, "INSERT INTO annees(label,actif) VALUES(?,0)", (label,))
    except: pass
    if d.get('dupliquer') and d.get('source'):
        src = d['source']
        if not fetchone(conn, "SELECT COUNT(*) as n FROM attributions WHERE annee=?", (label,))['n']:
            execute(conn, """INSERT INTO attributions(cours_id,classe_id,personnel_id,annee,groupe_num,heures_attr)
                             SELECT cours_id,classe_id,personnel_id,?,groupe_num,heures_attr
                             FROM attributions WHERE annee=?""", (label, src))
            execute(conn, """INSERT INTO titulaires(classe_id,personnel_id,annee)
                             SELECT classe_id,personnel_id,? FROM titulaires WHERE annee=?""", (label, src))
            execute(conn, """INSERT INTO ntpp_valeurs(categorie_id,annee,valeur)
                             SELECT categorie_id,?,valeur FROM ntpp_valeurs WHERE annee=?""", (label, src))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/annee/<path:label>', methods=['DELETE'])
@login_required
@admin_required
def del_annee(label):
    conn = get_db()
    r = fetchone(conn, "SELECT actif FROM annees WHERE label=?", (label,))
    if r and r['actif']:
        conn.close(); return jsonify({'ok':False,'error':"Impossible de supprimer l'année active"})
    for t in ('attributions','titulaires','eleves_options','ntpp_valeurs','mail_envois'):
        execute(conn, f"DELETE FROM {t} WHERE annee=?", (label,))
    execute(conn, "DELETE FROM annees WHERE label=?", (label,))
    conn.commit(); conn.close(); return jsonify({'ok':True})

# ── API Mail ──────────────────────────────────────────────────────────────────
@app.route('/api/mail/config', methods=['PUT'])
@login_required
@admin_required
def update_mail_config():
    d = request.json; conn = get_db()
    allowed = ('smtp_host','smtp_port','smtp_user','smtp_pass','from_name','signature')
    fields = [f"{k}=?" for k in d if k in allowed]
    vals   = [d[k] for k in d if k in allowed]
    if fields:
        execute(conn, f"UPDATE mail_config SET {','.join(fields)} WHERE id=1", vals)
        conn.commit()
    conn.close(); return jsonify({'ok':True})

@app.route('/api/mail/template', methods=['POST'])
@login_required
@admin_required
def save_template():
    d = request.json; conn = get_db()
    if d.get('id'):
        execute(conn, "UPDATE mail_templates SET nom=?,sujet=?,corps=? WHERE id=?",
                (d['nom'],d['sujet'],d['corps'],d['id']))
    else:
        execute(conn, "INSERT INTO mail_templates(nom,sujet,corps) VALUES(?,?,?)",
                (d['nom'],d['sujet'],d['corps']))
    conn.commit(); conn.close(); return jsonify({'ok':True})

@app.route('/api/mail/preview/<int:pid>')
@login_required
def mail_preview(pid):
    annee = request.args.get('annee', annee_active())
    tid   = request.args.get('template_id',1)
    conn  = get_db()
    p   = fetchone(conn, "SELECT * FROM personnel WHERE id=?", (pid,))
    t   = fetchone(conn, "SELECT * FROM mail_templates WHERE id=?", (tid,))
    cfg = fetchone(conn, "SELECT * FROM mail_config LIMIT 1")
    if not p or not t: conn.close(); return jsonify({'ok':False,'error':'Introuvable'})
    attrs = fetchall(conn, """
        SELECT f.nom as filiere, COALESCE(cl.nom,f.nom) as classe,
               c.nom, c.heures, c.type, a.groupe_num, a.heures_attr
        FROM attributions a JOIN cours c ON a.cours_id=c.id
        JOIN filieres f ON c.filiere_id=f.id LEFT JOIN classes cl ON a.classe_id=cl.id
        WHERE a.personnel_id=? AND a.annee=? ORDER BY f.ordre,c.type,c.ordre
    """, (pid, annee))
    titu = fetchall(conn, """
        SELECT f.nom as filiere, cl.nom as classe
        FROM titulaires t JOIN classes cl ON t.classe_id=cl.id
        JOIN filieres f ON cl.filiere_id=f.id WHERE t.personnel_id=? AND t.annee=?
    """, (pid, annee))
    conn.close()
    lines = [f"{'Filière':<20} {'Classe':<10} {'Cours':<35} {'Type':<6} {'H':>5}",'-'*80]
    total = 0
    for a in attrs:
        h = a['heures_attr'] if a['heures_attr'] else a['heures']
        total += h
        lines.append(f"{a['filiere']:<20} {a['classe']:<10} {a['nom']:<35} {a['type']:<6} {h:>4.0f}h")
    lines += ['-'*80, f"{'TOTAL':>72} {total:>4.0f}h"]
    titu_txt = ('Titulariats :\n' + '\n'.join(f"  - {t['filiere']} / {t['classe']}" for t in titu) + '\n\n') if titu else ''
    corps = t['corps'].replace('{prenom}', p['prenom'] or p['acronyme'])
    corps = corps.replace('{nom}', p['nom'] or '').replace('{acronyme}', p['acronyme'])
    corps = corps.replace('{annee}', annee).replace('{signature}', cfg.get('signature','') if cfg else '')
    corps = corps.replace('{titulariats}', titu_txt).replace('{tableau}', '\n'.join(lines))
    return jsonify({'ok':True,'sujet':t['sujet'].replace('{annee}',annee),'corps':corps,'email':p.get('email','')})

@app.route('/api/mail/send', methods=['POST'])
@login_required
@admin_required
def send_mail():
    d = request.json; pids = d.get('personnel_ids',[]); tid = d.get('template_id',1)
    annee = d.get('annee', annee_active())
    conn = get_db(); cfg = fetchone(conn, "SELECT * FROM mail_config LIMIT 1"); conn.close()
    if not cfg or not cfg.get('smtp_user') or not cfg.get('smtp_pass'):
        return jsonify({'ok':False,'error':'Configuration SMTP incomplète'})
    results = []
    for pid in pids:
        prev = mail_preview(pid)
        data = json.loads(prev.data)
        if not data.get('ok') or not data.get('email'):
            results.append({'id':pid,'ok':False,'error':'Email manquant'}); continue
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = data['sujet']
            msg['From']    = f"{cfg.get('from_name','')} <{cfg['smtp_user']}>"
            msg['To']      = data['email']
            msg.attach(MIMEText(data['corps'],'plain','utf-8'))
            with smtplib.SMTP_SSL(cfg['smtp_host'], int(cfg['smtp_port'])) as s:
                s.login(cfg['smtp_user'], cfg['smtp_pass']); s.send_message(msg)
            conn2 = get_db()
            execute(conn2, "INSERT INTO mail_envois(personnel_id,template_id,annee,date_envoi,statut) VALUES(?,?,?,?,?)",
                    (pid,tid,annee,datetime.now().isoformat(),'sent'))
            conn2.commit(); conn2.close()
            results.append({'id':pid,'ok':True})
        except Exception as e:
            conn2 = get_db()
            execute(conn2, "INSERT INTO mail_envois(personnel_id,template_id,annee,date_envoi,statut,erreur) VALUES(?,?,?,?,?,?)",
                    (pid,tid,annee,datetime.now().isoformat(),'error',str(e)))
            conn2.commit(); conn2.close()
            results.append({'id':pid,'ok':False,'error':str(e)})
    return jsonify({'ok':True,'results':results})

# ── Backup ────────────────────────────────────────────────────────────────────
@app.route('/api/backup')
@login_required
@admin_required
def backup():
    from database import USE_POSTGRES, SQLITE_PATH
    if USE_POSTGRES:
        return jsonify({'ok':False,'error':'Sauvegarde PostgreSQL non disponible via cette interface. Utilisez Railway dashboard.'})
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = SQLITE_PATH.replace('.db', f'_backup_{ts}.db')
    shutil.copy2(SQLITE_PATH, dst)
    return send_file(dst, download_name=f'attributions_backup_{ts}.db', as_attachment=True)

# ── Synthèse API ──────────────────────────────────────────────────────────────
@app.route('/api/synthese/<int:pid>')
@login_required
def synthese_detail(pid):
    annee = request.args.get('annee', annee_active())
    conn  = get_db()
    detail = fetchall(conn, """
        SELECT f.nom as filiere, COALESCE(cl.nom,f.nom) as classe,
               c.nom, c.type, a.groupe_num,
               CASE WHEN a.heures_attr IS NOT NULL THEN a.heures_attr ELSE c.heures END as h
        FROM attributions a JOIN cours c ON a.cours_id=c.id
        JOIN filieres f ON c.filiere_id=f.id LEFT JOIN classes cl ON a.classe_id=cl.id
        WHERE a.personnel_id=? AND a.annee=? ORDER BY f.ordre,c.type,c.ordre
    """, (pid, annee))
    titu = fetchall(conn, """
        SELECT f.nom as filiere, cl.nom as classe
        FROM titulaires t JOIN classes cl ON t.classe_id=cl.id
        JOIN filieres f ON cl.filiere_id=f.id WHERE t.personnel_id=? AND t.annee=?
    """, (pid, annee))
    conn.close()
    return jsonify({'attributions':detail,'titulariats':titu})

# ── Export Excel ──────────────────────────────────────────────────────────────
@app.route('/export/excel')
@login_required
def export_excel():
    import io; from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    annee = request.args.get('annee', annee_active())
    conn  = get_db()
    wb = Workbook(); wb.remove(wb.active)
    H = PatternFill('solid',fgColor='1F4E79'); HF = Font(bold=True,color='FFFFFF',name='Arial',size=10)
    HA = Alignment(horizontal='center',vertical='center')
    ws = wb.create_sheet('Données')
    headers = ['Filière','Classe','Cours','Type','Professeur','Groupe','Heures']
    for ci,h in enumerate(headers,1): c=ws.cell(1,ci,h); c.font=HF; c.fill=H; c.alignment=HA
    data = fetchall(conn, """
        SELECT f.nom,COALESCE(cl.nom,f.nom),c.nom,c.type,p.acronyme,a.groupe_num,
               CASE WHEN a.heures_attr IS NOT NULL THEN a.heures_attr ELSE c.heures END
        FROM attributions a JOIN cours c ON a.cours_id=c.id
        JOIN filieres f ON c.filiere_id=f.id JOIN personnel p ON a.personnel_id=p.id
        LEFT JOIN classes cl ON a.classe_id=cl.id
        WHERE a.annee=? ORDER BY f.ordre,c.type,c.ordre
    """, (annee,))
    EVEN = PatternFill('solid',fgColor='D6E4F0')
    for ri,row in enumerate(data,2):
        fill = EVEN if ri%2==0 else PatternFill('solid',fgColor='FFFFFF')
        for ci,v in enumerate(list(row.values()),1):
            cell=ws.cell(ri,ci,v); cell.fill=fill; cell.font=Font(name='Arial',size=9)
    conn.close()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=f'attributions_{annee}.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Initialisation au démarrage ───────────────────────────────────────────────
# Appelé par gunicorn ET python app.py
import atexit as _atexit

# DB initialized by gunicorn.conf.py on_starting hook

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    app.run(debug=False, port=5000)
